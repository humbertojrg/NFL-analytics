# Practice with excel

import openpyxl as xls
import matplotlib.pyplot as plt
import operator


wb = xls.load_workbook('NFL_Data.xlsx')
sheet_names = wb.get_sheet_names()

# Convert all sheet names from unicode to strings
for i in range(len(sheet_names)):
    sheet_names[i] = str(sheet_names[i])

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
    
# Function that takes in a sheet and creates a dictionary with opponents,
# points, and scores for each game

def totOffense(sheet1):
    totOff = dict() # create dictionary with team's name
    opps = Opponents(sheet1)
    totYards = totalYards(sheet1)
    points = pointsScored(sheet1)
    for i in range(len(opps)):
        totOff[opps[i]] = [defRank(opps[i]), points[i], totYards[i]] 
    return totOff


""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
# function to create a list of opponents
def Opponents(sheet1):
    opponents = [] # list of opponents
    col = 3 # column where opponents are found
    for i in range(1,sheet.max_row+1):
        value = str(sheet.cell(None,i,col).value) # value in cell as string
        if value in ['@','vs','None']: # takes out filler words and blanks
            continue
        else:
            opponents.append(value)
    return opponents

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""    
# function to create list of points scored in each game
def pointsScored(sheet1):
    points = [] # list of points scored
    col = 4 # column where points are found
    for i in range(1,sheet.max_row+1):
        result = str(sheet.cell(None,i,col).value) # value in cell as string, looking for game results
        if result in ['W', 'L', 'T']: # looks for Win, Loss, or Tie
            value = str(sheet.cell(None,i+1,col).value) # looks at value below which is the score
            score = getScore(result, value) # grabs the points scored from the cell
            points.append(score)
        else:
            continue
    return points
    
def getScore(result, value):
    hyp = value.find('-') #finds where the hyphen is which seperates the scores
    if result == 'T': 
        score = int(value[0:hyp])
    else:
        first_score = int(value[0:hyp])
        second_score = int(value[hyp+1:len(value)])
        if result == 'W': # team won so they got the higher score
            score = max(first_score,second_score)
        else:
            score = min(first_score,second_score)
    return score
     

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""        
  
def defRank(opp): # given a team, return the defensive rank of that team
    sheet_name = 'Def_Rank' # name of the sheet with the ranking info
    sheet = wb.get_sheet_by_name(sheet_name)
    tCol = 2
    rCol = 1
    teams = []
    ranks = []
    for i in range(1,sheet.max_row+1):
        teams.append(str(sheet.cell(None,i,tCol).value)) # creates list of teams
        ranks.append(str(sheet.cell(None,i,rCol).value)) # creates list of ranks
    indx = teams.index(opp) # find the index of the team
    rank = float(ranks[indx]) # uses index to find rank
    return rank
          
    
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
# function to create list of total yards in each game
def totalYards(sheet1):
    totYards = [] # list of total yards for each game
    for u in range(1,sheet.max_row+1):
        value = sheet.cell(None,u,5).value
        if type(value) == type(None):
            continue
        else:
            pYards = passingYards(u) # finds the passing yards
            rYards = rushingYards(u) # finds the rushing yards
            tot = pYards + rYards
            totYards.append(tot)
    return totYards                        
    
# function to find the passing yards in the cell
def passingYards(x):
    pCol = 5 # column where the passing yards are
    value = sheet.cell(None,x,pCol).value.replace(u'\xa0', ' ')
    value = str(value)
    space = value.find(' ')
    py = int(value[space+1:len(value)])
    return py 

# function to find the rushing yards in the cell
def rushingYards(x):
    rCol = 6 # column where the passing yards are
    value = sheet.cell(None,x,rCol).value.replace(u'\xa0', ' ')
    value = str(value)
    space = value.find(' ')
    ry = int(value[space+1:len(value)])
    return ry                 
    
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

# calculates the offensive score each team recieved when facing each opponent
def offScore(data):
    teams = data.keys() # list of the teams
    offensive_Scores_Data = dict()
    rInd = 0 # index of the opponents rank
    pInd = 1 # index of the points scored
    yInd = 2 # index of the yards gained offensively 
    for team in teams:
        offensive_Scores_Data[team]=dict()
        opponents = data[team].keys()
        for opp in opponents:
            offensive_Scores_Data[team][opp] = calcOffScore(data[team][opp][rInd], data[team][opp][pInd], data[team][opp][yInd])
    return offensive_Scores_Data       

def calcOffScore(rank, points, yards):
    score = round((rank) * (max((points/7),1) + (yards/100)),5)
    return score


""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""    
# add all the offensive scores from the season
def add_Off_Scores(data):
    teams = data.keys() # list of the teams
    off_Scores = dict()
    for team in teams:
        opponents = data[team].keys()
        off_Scores[team] = 0
        for opp in opponents:
            off_Scores[team] += round(data[team][opp],5) # adds up all the scores from the season
    return off_Scores

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""" 
# given the teams and thier combined score, finds the max aka Top Offensive Team
def max_Score(data):
     scores = list(data.values())
     teams = list(data.keys())
     return teams[scores.index(max(scores))]
                   

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""" 
# rounds the overall scores for each team
def round_Scores(data):
     teams = data.keys()
     new_Scores = dict()
     for team in teams:
         new_Scores[team] = round(data[team]/10,6)
     return new_Scores
         

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""" 
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

# creates a dictionary of all the offensive data for each team and each team they faced
team_Offensive_Data = dict()
for i in range(2,len(sheet_names)):
    sheet = wb.get_sheet_by_name(sheet_names[i])
    team_Offensive_Data[sheet_names[i]] = totOffense(sheet)
    
    
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

  
### OFFENSIVE RANKINGS ###
    
offensive_Scores = offScore(team_Offensive_Data) 
overall_Scores = round_Scores(add_Off_Scores(offensive_Scores))
top_Rank = max_Score(overall_Scores)
sorted_Scores = sorted(overall_Scores.items(), key=operator.itemgetter(1))
sorted_Scores.reverse()
    
            
                   
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""" 
### DEFENSIVE DATA ###


def construct_Def_Data(data):
    defensive_Data = dict()
    teams = data.keys()
    for team in teams:
        defensive_Data[team]=dict()
        opponents = data[team].keys()
        for opp in opponents:
            defensive_Data[team][opp] = [data[opp][team][1],data[opp][team][2]] 
    return defensive_Data
      
team_Defensive_Data = construct_Def_Data(team_Offensive_Data)    


"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""" 

def list_Off_Yards(data):
    offYardage = dict()
    teams = data.keys()
    for team in teams:
        offYardage[team]=dict()
        opponents = data[team].keys()
        for opp in opponents:
           offYardage[team][opp] = data[team][opp][2]
    
    return offYardage

off_Yards = list_Off_Yards(team_Offensive_Data)
     
        
def list_Off_Points(data):
    offPoints = dict()
    teams = data.keys()
    for team in teams:
        offPoints[team]=dict()
        opponents = data[team].keys()
        for opp in opponents:
           offPoints[team][opp] = data[team][opp][1]
    
    return offPoints

off_Points = list_Off_Points(team_Offensive_Data)        
        
        
        
        
        




