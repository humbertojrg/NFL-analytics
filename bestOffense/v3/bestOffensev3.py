""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"Given the offensive stats for all teams in the NFL, we will look at how each"
"team performed against each team they faced in terms of points scored in"
"relation to the opposing teams normal distribution of points scored agaisnt"
"them"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""


""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"Importing all neccessary modules"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

import openpyxl as xls
import numpy as np
import matplotlib.pyplot as plt

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"Importing Data from the Excel File"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

wb = xls.load_workbook('NFL_Data2.xlsx')
NFLteams = wb.get_sheet_names()

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"OFFENSIVE DATA"
"Reading the Data from the Excel File and storing them into variables,"
"This is the points data meaning that the dictionary created will consist of"
"teams, the team they faced, and the points scored against that team"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

# Function that takes in a sheet and creates a dictionary with opponents,
# and the points scored against them

def MatchupsDict(team):
    sheet =  sheet = wb.get_sheet_by_name(team)
    totPoints = dict() # create dictionary with team's name
    opps = Opponents(sheet)
    points = pointsScored(sheet)
    for i in range(len(opps)):
        totPoints[opps[i]] = points[i]
    return totPoints

# function to create a list of opponents a team has faced
def Opponents(sheet1):
    opponents = [] # list of opponents
    col = 3 # column where opponents are found
    for i in range(1,sheet1.max_row+1):
        value = str(sheet1.cell(None,i,col).value) # value in cell as string
        if value in ['@','vs','None']: # takes out filler words and blanks
            continue
        else:
            opponents.append(value)
    return opponents

# function to create list of points scored in each game by the team being 
# looked at
def pointsScored(sheet1):
    points = [] # list of points scored
    col = 4 # column where points are found
    for i in range(1,sheet1.max_row+1):
        result = str(sheet1.cell(None,i,col).value) # value in cell as string, looking for game results
        if result in ['W', 'L', 'T']: # looks for Win, Loss, or Tie
            value = str(sheet1.cell(None,i+1,col).value) # looks at value below which is the score
            score = getScore(result, value) # grabs the points scored from the cell
            points.append(score)
        else:
            continue
    return points
  
# depending on the result of the game, assigns either the winning or losing score
def getScore(result, value):
    hyp = value.find('-') #finds where the hyphen is which seperates the scores
    if result == 'T': 
        score = int(value[0:hyp])
    else:
        first_score = int(value[0:hyp])
        second_score = int(value[hyp+1:len(value)])
        if result == 'W': # team won so they got the higher score
            score = max(first_score,second_score)
        else:
            score = min(first_score,second_score)
    return score


# MAIN function that creates a dictionary where each team has its own dictionary
# containing each team they have faced and the points scored against that team  

def createPointsDict():
    masterdict = {}
    for team in NFLteams:
        masterdict[team] = MatchupsDict(team)
    return masterdict               
    
offPointsData = createPointsDict()



""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"DEFENSIVE DATA"
"Reading the Data from the Excel File and storing them into variables,"
"This is the defensive points data meaning that the dictionary created will" 
"consist of teams, the team they faced, and the points allowed against" 
"that team"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
# using the offensive data, creates a dictionary for points allowed
# by a team for each game they played
def construct_Def_Data(data):
    defensive_Data = dict()
    teams = data.keys()
    for team in teams:
        defensive_Data[team]=dict()
        opponents = data[team].keys()
        for opp in opponents:
            defensive_Data[team][opp] = data[opp][team]
    return defensive_Data
      
defPointsData = construct_Def_Data(offPointsData)  


""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"Given a teams points allowed during a season, we calculate the total points"
"allowed, the average per game, and the standard deviation"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

# given a dictionary, removes the teams the faced and used just a list of 
# values in the dictionary
def createListofPoints(data):
    ndict = {}
    teams = data.keys()
    for team in teams:
        ndict[team] = list(data[team].values())
    
    return ndict

defPointsDataList = createListofPoints(defPointsData)

# takes in the data comprising of list of values, finds the avg and std 
def NdistrOfPoints(data):
    ndict = {}
    teams = data.keys()
    for team in teams:
        avg = np.mean(data[team])
        stdev = np.std(data[team])
        ndict[team] = [avg,stdev]
    return ndict
    

defStats = NdistrOfPoints(defPointsDataList)



""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
"Given how many points each team scored against their opponent, and what each"
"opponents average points allowed and standard deviation of points allowed,"
"we can gauge how well each team did agaisnt each opponent they faced"
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""


# compares how each team did against each opponent by comparing how many points
# they scored in relation to their opponents average points allowed and 
# the standard deviation from that average
def compare(offdata,defstats):
    newdict = {}
    teams = offdata.keys()
    for team in teams:
        opponents = offdata[team].keys()
        newdict[team] = {}
        for opp in opponents:
            pointsScored = offdata[team][opp]
            defavg = defstats[opp][0]
            defstd = defstats[opp][1]
            score = ((pointsScored-defavg)*1.0)/defstd
            newdict[team][opp] = score
    return newdict

offStatsComparison = compare(offPointsData,defStats)

offStatsList = createListofPoints(offStatsComparison)


# Given how each team did against each opponent, the scores are added up 
# to create a total
def combineScores(data):
    newdict = {}
    teams = data.keys()
    for team in teams:
        newdict[team] = sum(data[team])
    return newdict


# Results of the analysis
overallOffResults = [(v, k) for k, v in combineScores(offStatsList).items()]
overallOffResults.sort(reverse=True)































